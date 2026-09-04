from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, Response
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import json, os, csv, io, re, base64
from pypdf import PdfReader
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors as rl_colors
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics import renderPDF
from psycopg2cffi import compat
compat.register()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'eps-ibt-portal-secret-key')
app.jinja_env.filters['from_json'] = json.loads
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
    'DATABASE_URL', 'sqlite:///database.db'
).replace('postgres://', 'postgresql://')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
}
db = SQLAlchemy(app)

SUBJECTS  = ['English', 'Mathematics', 'Science', 'Reasoning']
DT_SUBJECTS = ['English', 'Hindi', 'Maths', 'Science', 'Urdu', 'ICT']

# ── IB CONSTANTS ──────────────────────────────────────────────────────────────
TERMS = ['Term 1', 'Term 2', 'Term 3']
RATING_SCALE = {1: 'Beginning', 2: 'Developing', 3: 'Achieved', 4: 'Exceeding'}
RATING_COLORS = {1: '#ef4444', 2: '#f59e0b', 3: '#3b82f6', 4: '#10b981'}

LEARNER_PROFILE = [
    ('Inquirer',      'Nurtures curiosity and love of learning. Acquires skills to inquire and research independently.'),
    ('Knowledgeable', 'Explores concepts across disciplines. Engages with local and global issues.'),
    ('Thinker',       'Uses critical and creative thinking to tackle complex problems and make ethical decisions.'),
    ('Communicator',  'Expresses ideas confidently in multiple modes and collaborates effectively.'),
    ('Principled',    'Acts with integrity and honesty. Takes responsibility for actions and consequences.'),
    ('Open-minded',   'Appreciates own and others cultures. Seeks and evaluates diverse perspectives.'),
    ('Caring',        'Shows empathy, compassion and respect. Makes a positive difference to others.'),
    ('Risk-taker',    'Approaches uncertainty with courage and creativity. Explores new ideas and strategies.'),
    ('Balanced',      'Understands importance of intellectual, physical and emotional balance.'),
    ('Reflective',    'Thoughtfully considers learning and experiences to improve understanding and growth.'),
]

ATL_SKILLS = {
    'Communication': {
        'Grade 3': ['Shares ideas clearly in class discussions','Listens attentively when others speak','Writes sentences with a clear beginning and end','Reads aloud with expression and understanding'],
        'Grade 4': ['Expresses ideas using appropriate vocabulary','Asks clarifying questions during discussions','Organises written work with introduction and conclusion','Reads and summarises key information'],
        'Grade 5': ['Communicates effectively for different audiences','Evaluates and responds to others ideas respectfully','Writes structured arguments with evidence','Analyses texts and identifies authors purpose'],
    },
    'Self-Management': {
        'Grade 3': ['Brings required materials to class','Follows classroom routines independently','Manages time during tasks with teacher support','Stays on task with minimal reminders'],
        'Grade 4': ['Organises work and materials independently','Sets personal goals with teacher guidance','Manages time effectively during class activities','Reflects on learning with prompting'],
        'Grade 5': ['Plans and organises multi-step tasks independently','Sets and monitors personal learning goals','Manages time across multiple responsibilities','Reflects critically on own learning and progress'],
    },
    'Research': {
        'Grade 3': ['Identifies information from given sources','Distinguishes between fact and opinion with support','Records information in own words','Uses library and digital resources with guidance'],
        'Grade 4': ['Selects relevant information from multiple sources','Identifies reliable vs unreliable sources','Organises research notes effectively','Cites sources with teacher support'],
        'Grade 5': ['Independently researches using varied sources','Evaluates credibility and bias in sources','Synthesises information to form conclusions','Properly attributes sources and avoids plagiarism'],
    },
    'Thinking': {
        'Grade 3': ['Makes connections between new and prior learning','Identifies patterns and sequences','Generates ideas during brainstorming','Solves simple problems with support'],
        'Grade 4': ['Asks why and what if questions','Applies knowledge to new situations','Identifies cause and effect relationships','Evaluates solutions to problems'],
        'Grade 5': ['Analyses information from multiple perspectives','Creates original solutions to complex problems','Transfers learning across subjects','Justifies reasoning with evidence'],
    },
    'Social': {
        'Grade 3': ['Takes turns and shares during group activities','Shows kindness and respect to classmates','Accepts different roles in group work','Resolves conflicts with teacher support'],
        'Grade 4': ['Contributes meaningfully to group discussions','Encourages and supports peers','Adapts role based on group needs','Resolves minor conflicts independently'],
        'Grade 5': ['Leads and participates effectively in groups','Considers diverse perspectives in collaboration','Negotiates and compromises to achieve goals','Mediates conflicts constructively'],
    },
}
GRADES    = ['Grade 3', 'Grade 4', 'Grade 5']
SECTIONS_BY_SUBJECT = {
    'English':     ['Reading Comprehension', 'Grammar', 'Spelling', 'Vocabulary', 'Punctuation'],
    'Mathematics': ['Number Operations', 'Algebra', 'Geometry', 'Measurement', 'Data & Statistics'],
    'Science':     ['Life Science', 'Physical Science', 'Earth Science', 'Scientific Inquiry'],
    'Reasoning':   ['Verbal Reasoning', 'Non-Verbal Reasoning', 'Logical Thinking', 'Pattern Recognition'],
}

# ── MODELS ───────────────────────────────────────────────────────────────────

class User(db.Model):
    id       = db.Column(db.Integer, primary_key=True)
    name     = db.Column(db.String(100), nullable=False)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role     = db.Column(db.String(20), nullable=False)
    grade    = db.Column(db.String(20), nullable=True)
    section  = db.Column(db.String(10), nullable=True)
    created  = db.Column(db.DateTime, default=datetime.utcnow)
    results  = db.relationship('TestResult', backref='student', lazy=True, cascade='all,delete-orphan')

class MockTest(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(200), nullable=False)
    subject    = db.Column(db.String(50), nullable=False)
    grade      = db.Column(db.String(20), nullable=False)
    difficulty = db.Column(db.String(20), default='Medium')
    duration   = db.Column(db.Integer, default=40)
    status     = db.Column(db.String(10), default='draft')
    questions  = db.Column(db.Text, default='[]')
    created    = db.Column(db.DateTime, default=datetime.utcnow)
    results    = db.relationship('TestResult', backref='test', lazy=True, cascade='all,delete-orphan')

class TestResult(db.Model):
    id             = db.Column(db.Integer, primary_key=True)
    student_id     = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    test_id        = db.Column(db.Integer, db.ForeignKey('mock_test.id'), nullable=False)
    score          = db.Column(db.Integer, default=0)
    total          = db.Column(db.Integer, default=0)
    percent        = db.Column(db.Float, default=0.0)
    answers        = db.Column(db.Text, default='{}')
    section_scores = db.Column(db.Text, default='{}')
    time_taken     = db.Column(db.Integer, default=0)
    taken_at       = db.Column(db.DateTime, default=datetime.utcnow)

# ── IB HOLISTIC MODELS ───────────────────────────────────────────────────────

ATL_SKILLS = {
    'Communication':    ['Reading & Writing', 'Listening & Speaking', 'Presenting Ideas'],
    'Self-Management':  ['Organisation', 'Time Management', 'Reflection'],
    'Research':         ['Information Literacy', 'Media Literacy', 'Using Sources'],
    'Thinking':         ['Critical Thinking', 'Creative Thinking', 'Problem Solving'],
    'Social':           ['Collaboration', 'Respecting Others', 'Leadership'],
}

LEARNER_PROFILE = [
    'Inquirer', 'Knowledgeable', 'Thinker', 'Communicator', 'Principled',
    'Open-minded', 'Caring', 'Risk-taker', 'Balanced', 'Reflective'
]

LP_DESCRIPTIONS = {
    'Inquirer':      'Nurtures curiosity and love of learning',
    'Knowledgeable': 'Explores concepts across disciplines',
    'Thinker':       'Uses critical and creative thinking skills',
    'Communicator':  'Expresses ideas confidently in many ways',
    'Principled':    'Acts with integrity and honesty',
    'Open-minded':   'Appreciates other perspectives and cultures',
    'Caring':        'Shows empathy, compassion and respect',
    'Risk-taker':    'Approaches uncertainty with courage',
    'Balanced':      'Understands importance of all aspects of life',
    'Reflective':    'Thoughtfully considers own learning and growth',
}

ATL_DESCRIPTORS = {
    'Grade 3': {
        1: 'I am just starting to learn this skill',
        2: 'I sometimes use this skill with help',
        3: 'I often use this skill on my own',
        4: 'I always use this skill and help my friends',
    },
    'Grade 4': {
        1: 'I am aware of this skill but need reminders',
        2: 'I use this skill sometimes with support',
        3: 'I independently use this skill most of the time',
        4: 'I consistently use and model this skill for others',
    },
    'Grade 5': {
        1: 'I am beginning to develop this skill with guidance',
        2: 'I apply this skill in familiar situations with some support',
        3: 'I consistently apply this skill across different contexts',
        4: 'I demonstrate this skill with sophistication and mentor peers',
    },
}

LP_DESCRIPTORS = {
    'Grade 3': {1:'Just starting',  2:'Sometimes',  3:'Often',        4:'Always'},
    'Grade 4': {1:'Beginning',      2:'Developing', 3:'Achieved',     4:'Exceeding'},
    'Grade 5': {1:'Beginning',      2:'Developing', 3:'Consistently', 4:'Exemplary'},
}

TERMS = ['Term 1', 'Term 2', 'Term 3']


class ATLRating(db.Model):
    __tablename__ = 'atl_rating'
    id           = db.Column(db.Integer, primary_key=True)
    student_id   = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    rater_id     = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    rater_type   = db.Column(db.String(10), nullable=False)  # 'teacher' or 'student'
    term         = db.Column(db.String(20), nullable=False)
    category     = db.Column(db.String(50), nullable=False)
    skill        = db.Column(db.String(100), nullable=False)
    rating       = db.Column(db.Integer, nullable=False)  # 1-4
    comment      = db.Column(db.Text, nullable=True)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)
    student      = db.relationship('User', foreign_keys=[student_id])
    rater        = db.relationship('User', foreign_keys=[rater_id])


class LearnerProfileRating(db.Model):
    __tablename__ = 'lp_rating'
    id           = db.Column(db.Integer, primary_key=True)
    student_id   = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    rater_id     = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    rater_type   = db.Column(db.String(10), nullable=False)  # 'teacher' or 'student'
    term         = db.Column(db.String(20), nullable=False)
    attribute    = db.Column(db.String(50), nullable=False)
    rating       = db.Column(db.Integer, nullable=False)  # 1-4
    evidence     = db.Column(db.Text, nullable=True)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)
    student      = db.relationship('User', foreign_keys=[student_id])
    rater        = db.relationship('User', foreign_keys=[rater_id])


class StudentReflection(db.Model):
    __tablename__ = 'student_reflection'
    id           = db.Column(db.Integer, primary_key=True)
    student_id   = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    term         = db.Column(db.String(20), nullable=False)
    reflection   = db.Column(db.Text, nullable=False)
    goals        = db.Column(db.Text, nullable=True)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)
    student      = db.relationship('User', foreign_keys=[student_id])


# ── IB HOLISTIC MODELS ───────────────────────────────────────────────────────

ATL_SKILLS = {
    'Communication':   ['Exchanges information clearly','Listens actively','Reads with understanding','Writes effectively','Presents ideas confidently'],
    'Self-Management': ['Organises time and tasks','Manages materials','Reflects on learning','Shows perseverance','Sets goals'],
    'Research':        ['Finds reliable information','Evaluates sources','Documents information','Uses data effectively'],
    'Thinking':        ['Thinks critically','Thinks creatively','Transfers knowledge','Solves problems','Makes connections'],
    'Social':          ['Collaborates respectfully','Supports peers','Manages disagreements','Contributes to group'],
}

ATL_DESCRIPTORS = {
    'Grade 3': {
        1: 'Beginning — Needs a lot of support to demonstrate this skill',
        2: 'Developing — Demonstrates this skill with some support',
        3: 'Achieved — Demonstrates this skill independently most of the time',
        4: 'Exceeding — Consistently demonstrates this skill and supports others',
    },
    'Grade 4': {
        1: 'Beginning — Rarely demonstrates this skill even with support',
        2: 'Developing — Demonstrates this skill with teacher guidance',
        3: 'Achieved — Independently demonstrates this skill consistently',
        4: 'Exceeding — Applies this skill in new contexts and models it for peers',
    },
    'Grade 5': {
        1: 'Beginning — Limited evidence of this skill',
        2: 'Developing — Growing evidence with scaffolding required',
        3: 'Achieved — Consistent and independent application of skill',
        4: 'Exceeding — Sophisticated application; leads and mentors others',
    },
}

LEARNER_PROFILE = [
    ('Inquirer',     '🔍', 'Nurtures curiosity and love of learning'),
    ('Knowledgeable','📚', 'Develops conceptual understanding across disciplines'),
    ('Thinker',      '💡', 'Applies critical and creative thinking skills'),
    ('Communicator', '🗣️', 'Expresses ideas confidently in multiple modes'),
    ('Principled',   '⚖️', 'Acts with integrity, honesty and strong ethics'),
    ('Open-minded',  '🌍', 'Appreciates perspectives, cultures and ideas'),
    ('Caring',       '❤️', 'Shows empathy, compassion and respect for others'),
    ('Risk-taker',   '🚀', 'Approaches uncertainty with courage and forethought'),
    ('Balanced',     '⚖️', 'Understands importance of balance in all life areas'),
    ('Reflective',   '🪞', 'Thoughtfully considers learning and personal growth'),
]

TERMS = ['Term 1', 'Term 2', 'Term 3']


# ── IB DEVELOPMENT MODELS ────────────────────────────────────────────────────

# ATL descriptors per grade and skill
ATL_SKILLS = {
    'Communication': {
        'Grade 3': ['Listens attentively during discussions', 'Expresses ideas clearly in writing', 'Uses appropriate vocabulary', 'Participates in group activities'],
        'Grade 4': ['Organises ideas logically in writing', 'Listens and responds thoughtfully', 'Uses subject-specific vocabulary', 'Presents information to an audience'],
        'Grade 5': ['Communicates complex ideas persuasively', 'Adapts communication style for audience', 'Uses evidence to support arguments', 'Gives and receives constructive feedback'],
    },
    'Self-Management': {
        'Grade 3': ['Follows classroom routines independently', 'Completes tasks within given time', 'Keeps materials organised', 'Manages emotions appropriately'],
        'Grade 4': ['Sets personal learning goals', 'Plans and prioritises tasks', 'Reflects on own learning', 'Manages distractions effectively'],
        'Grade 5': ['Develops strategies to overcome challenges', 'Monitors own progress toward goals', 'Demonstrates perseverance', 'Balances responsibilities effectively'],
    },
    'Research': {
        'Grade 3': ['Asks relevant questions', 'Locates information from given sources', 'Records findings in own words', 'Identifies fact vs. opinion'],
        'Grade 4': ['Selects appropriate sources', 'Evaluates reliability of information', 'Takes organised notes', 'Cites sources used'],
        'Grade 5': ['Formulates focused research questions', 'Critically evaluates multiple sources', 'Synthesises information effectively', 'Acknowledges intellectual property'],
    },
    'Thinking': {
        'Grade 3': ['Makes connections between ideas', 'Identifies patterns and relationships', 'Offers creative solutions', 'Asks "why" and "what if" questions'],
        'Grade 4': ['Analyses cause and effect', 'Evaluates different perspectives', 'Applies knowledge to new situations', 'Justifies opinions with reasons'],
        'Grade 5': ['Evaluates arguments critically', 'Generates and tests hypotheses', 'Transfers learning across subjects', 'Reflects on thinking processes (metacognition)'],
    },
    'Social': {
        'Grade 3': ['Takes turns and shares fairly', 'Respects different opinions', 'Helps classmates when needed', 'Follows agreed group rules'],
        'Grade 4': ['Contributes meaningfully to group work', 'Resolves conflicts respectfully', 'Encourages others', 'Takes on different group roles'],
        'Grade 5': ['Leads and supports team efforts', 'Negotiates and compromises effectively', 'Advocates for others', "Builds on others ideas constructively"],
    },
}

IB_LEARNER_PROFILE = [
    ('Inquirer',     '🔍', 'Nurtures curiosity, developing skills for inquiry and research'),
    ('Knowledgeable','📚', 'Explores concepts across disciplines with depth and breadth'),
    ('Thinker',      '💡', 'Uses critical and creative thinking to tackle complex problems'),
    ('Communicator', '🗣️', 'Expresses ideas confidently in more than one language'),
    ('Principled',   '⚖️', 'Acts with integrity, honesty and a strong sense of fairness'),
    ('Open-minded',  '🌍', 'Appreciates own culture and perspectives of others'),
    ('Caring',       '❤️', 'Shows empathy, compassion and respect for others'),
    ('Risk-taker',   '🚀', 'Approaches uncertainty with courage and forethought'),
    ('Balanced',     '⚖️', 'Understands importance of physical, mental and emotional balance'),
    ('Reflective',   '🪞', 'Thoughtfully considers learning and personal development'),
]

RATING_LABELS = {1: 'Beginning', 2: 'Developing', 3: 'Achieved', 4: 'Exceeding'}
TERMS = ['Term 1', 'Term 2', 'Term 3']


# ── IB DEVELOPMENT MODELS ────────────────────────────────────────────────────

ATL_CATEGORIES = {
    'Communication':   ['Reading & Writing', 'Listening & Speaking', 'Presenting Ideas'],
    'Self-Management': ['Organisation', 'Time Management', 'Emotional Regulation'],
    'Research':        ['Finding Information', 'Evaluating Sources', 'Recording Data'],
    'Thinking':        ['Critical Thinking', 'Creative Thinking', 'Problem Solving'],
    'Social':          ['Collaboration', 'Respecting Others', 'Conflict Resolution'],
}

LEARNER_PROFILE = [
    ('Inquirer',      '🔍', 'Develops curiosity and love of learning'),
    ('Knowledgeable', '📚', 'Explores concepts across disciplines'),
    ('Thinker',       '💡', 'Applies critical and creative thinking'),
    ('Communicator',  '🗣️', 'Expresses ideas confidently'),
    ('Principled',    '⚖️', 'Acts with integrity and honesty'),
    ('Open-minded',   '🌍', 'Appreciates other perspectives'),
    ('Caring',        '❤️', 'Shows empathy and compassion'),
    ('Risk-taker',    '🚀', 'Approaches uncertainty with courage'),
    ('Balanced',      '⚡', 'Balances intellectual and personal growth'),
    ('Reflective',    '🪞', 'Thoughtfully considers own learning'),
]

ATL_DESCRIPTORS = {
    'Grade 3': {
        1: 'Beginning — Needs significant support',
        2: 'Developing — Shows some understanding with guidance',
        3: 'Achieved — Demonstrates skill independently',
        4: 'Exceeding — Models skill and supports peers',
    },
    'Grade 4': {
        1: 'Beginning — Rarely demonstrates the skill',
        2: 'Developing — Sometimes demonstrates with prompting',
        3: 'Achieved — Consistently demonstrates the skill',
        4: 'Exceeding — Extends and applies skill in new contexts',
    },
    'Grade 5': {
        1: 'Beginning — Limited awareness of the skill',
        2: 'Developing — Growing awareness, inconsistent application',
        3: 'Achieved — Confident and consistent application',
        4: 'Exceeding — Exemplary — leads and inspires others',
    },
}

TERMS = ['Term 1', 'Term 2', 'Term 3']
ACADEMIC_YEAR = '2025-26'


# ── IB DEVELOPMENT MODELS ────────────────────────────────────────────────────

TERMS = ['Term 1', 'Term 2', 'Term 3']

LEARNER_PROFILE = [
    ('inquirer',    '🔍 Inquirer',     'Nurtures curiosity, develops skills for inquiry and research'),
    ('knowledgeable','📚 Knowledgeable','Develops and uses conceptual understanding across disciplines'),
    ('thinker',     '💭 Thinker',      'Uses critical and creative thinking to analyse complex problems'),
    ('communicator','💬 Communicator', 'Expresses confidently in more than one language and in many ways'),
    ('principled',  '⚖️ Principled',   'Acts with integrity and honesty, with strong sense of fairness'),
    ('open_minded', '🌍 Open-minded',  'Appreciates own culture and others, seeks diverse perspectives'),
    ('caring',      '❤️ Caring',       'Shows empathy, compassion and respect, makes a positive difference'),
    ('risk_taker',  '🚀 Risk-taker',   'Approaches uncertainty with forethought and determination'),
    ('balanced',    '⚖️ Balanced',     'Understands importance of intellectual, physical and emotional balance'),
    ('reflective',  '🪞 Reflective',   'Thoughtfully considers the world and own ideas and experience'),
]

ATL_CATEGORIES = {
    'communication': {
        'label': '💬 Communication',
        'skills': ['Reading & Writing', 'Listening & Speaking', 'Digital Communication'],
        'descriptors': {
            'Grade 3': {
                1: 'Rarely communicates ideas clearly',
                2: 'Sometimes shares ideas with support',
                3: 'Usually communicates ideas clearly',
                4: 'Always communicates ideas clearly and confidently',
            },
            'Grade 4': {
                1: 'Struggles to express ideas in writing or speech',
                2: 'Expresses basic ideas with some support',
                3: 'Expresses ideas clearly in writing and speech',
                4: 'Expresses ideas with clarity, detail and confidence',
            },
            'Grade 5': {
                1: 'Has difficulty structuring ideas for different audiences',
                2: 'Communicates ideas with inconsistent structure',
                3: 'Communicates effectively for purpose and audience',
                4: 'Communicates with sophistication adapting to any context',
            },
        }
    },
    'self_management': {
        'label': '🗂️ Self-Management',
        'skills': ['Organisation', 'Time Management', 'Emotional Regulation'],
        'descriptors': {
            'Grade 3': {
                1: 'Needs constant reminders to stay organised',
                2: 'Sometimes organises tasks with reminders',
                3: 'Usually organises work and meets deadlines',
                4: 'Consistently self-organised and meets all deadlines',
            },
            'Grade 4': {
                1: 'Rarely manages time or materials independently',
                2: 'Manages time with frequent teacher support',
                3: 'Manages time and materials independently',
                4: 'Excellently manages time, materials and emotions',
            },
            'Grade 5': {
                1: 'Struggles to set goals or monitor own learning',
                2: 'Sets goals but rarely follows through consistently',
                3: 'Sets and works toward goals with reflection',
                4: 'Independently sets, monitors and achieves goals',
            },
        }
    },
    'research': {
        'label': '🔎 Research',
        'skills': ['Information Literacy', 'Media Literacy', 'Note-taking'],
        'descriptors': {
            'Grade 3': {
                1: 'Rarely finds or uses information independently',
                2: 'Finds information with significant guidance',
                3: 'Finds and uses relevant information with some guidance',
                4: 'Independently researches and evaluates information',
            },
            'Grade 4': {
                1: 'Cannot distinguish reliable from unreliable sources',
                2: 'Sometimes identifies reliable sources with help',
                3: 'Usually selects reliable and relevant sources',
                4: 'Consistently evaluates and synthesises quality sources',
            },
            'Grade 5': {
                1: 'Rarely questions or evaluates information sources',
                2: 'Begins to question sources with teacher prompting',
                3: 'Questions sources and identifies bias',
                4: 'Critically evaluates sources for bias, purpose and reliability',
            },
        }
    },
    'thinking': {
        'label': '💡 Thinking',
        'skills': ['Critical Thinking', 'Creative Thinking', 'Transfer'],
        'descriptors': {
            'Grade 3': {
                1: 'Rarely applies prior knowledge to new tasks',
                2: 'Sometimes connects ideas with teacher support',
                3: 'Often makes connections between ideas',
                4: 'Consistently applies thinking across contexts',
            },
            'Grade 4': {
                1: 'Struggles to generate ideas or solutions independently',
                2: 'Generates basic ideas with guidance',
                3: 'Generates creative and logical ideas independently',
                4: 'Consistently generates innovative, well-reasoned ideas',
            },
            'Grade 5': {
                1: 'Rarely analyses or evaluates complex problems',
                2: 'Begins to analyse problems with support',
                3: 'Analyses problems and proposes reasoned solutions',
                4: 'Critically analyses and evaluates with depth and nuance',
            },
        }
    },
    'social': {
        'label': '🤝 Social',
        'skills': ['Collaboration', 'Conflict Resolution', 'Respect for Others'],
        'descriptors': {
            'Grade 3': {
                1: 'Rarely works cooperatively in groups',
                2: 'Sometimes cooperates with reminders',
                3: 'Usually cooperates and contributes to group work',
                4: 'Always contributes positively and supports peers',
            },
            'Grade 4': {
                1: "Struggles to listen to or consider others views",
                2: 'Sometimes listens and considers others with support',
                3: 'Listens actively and considers diverse perspectives',
                4: 'Champions inclusion and mediates group conflict positively',
            },
            'Grade 5': {
                1: 'Rarely takes shared responsibility in collaborative tasks',
                2: 'Takes limited responsibility in collaborative tasks',
                3: 'Takes responsibility and supports team goals',
                4: 'Leads and inspires effective collaboration',
            },
        }
    },
}

RATING_LABELS = {1: 'Beginning', 2: 'Developing', 3: 'Achieved', 4: 'Exceeding'}
RATING_COLORS = {1: '#fee2e2', 2: '#fef9c3', 3: '#dcfce7', 4: '#dbeafe'}
RATING_TEXT   = {1: '#dc2626', 2: '#92400e', 3: '#166534', 4: '#1d4ed8'}


# ── IB HOLISTIC MODELS ───────────────────────────────────────────────────────

# ── HELPERS ──────────────────────────────────────────────────────────────────

def login_required(role=None):
    from functools import wraps
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            if role and session.get('role') != role:
                flash('Access denied.', 'error')
                return redirect(url_for('login'))
            return f(*args, **kwargs)
        return decorated
    return decorator

def safe_avg(lst):
    lst = [x for x in lst if x is not None]
    return round(sum(lst)/len(lst), 1) if lst else 0

def generate_username(name, grade):
    first = re.sub(r'[^a-z0-9]', '', name.split()[0].lower())
    grade_num = re.sub(r'[^0-9]', '', grade)
    base = f"{first}_g{grade_num}"
    username = base
    counter = 1
    while User.query.filter_by(username=username).first():
        username = f"{base}{counter}"
        counter += 1
    return username

def generate_password(name, grade):
    first3 = name[:3].capitalize()
    grade_num = re.sub(r'[^0-9]', '', grade)
    return f"EPS@{first3}{grade_num}"

def parse_pdf_questions(file_stream):
    reader = PdfReader(file_stream)
    text = ""
    for page in reader.pages:
        text += page.extract_text() + "\n"
    text = re.sub(r'  +', ' ', text)
    answers = {}
    ak_match = re.search(r'Answer\s+Key.*?(?=\Z)', text, re.DOTALL | re.IGNORECASE)
    if ak_match:
        ak_text = ak_match.group()
        pairs = re.findall(r'(\d+)\s+([A-D])', ak_text)
        for qnum, ans in pairs:
            answers[int(qnum)] = ans
        text = text[:ak_match.start()]
    question_blocks = re.split(r'\n(?=\d+\.\s)', text)
    questions = []
    for block in question_blocks:
        block = block.strip()
        num_match = re.match(r'^(\d+)\.\s+(.+)', block, re.DOTALL)
        if not num_match:
            continue
        qnum = int(num_match.group(1))
        rest = num_match.group(2).strip()
        option_pattern = r'([A-D])\)\s+(.+?)(?=\s+[A-D]\)\s+|\Z)'
        options_found = re.findall(option_pattern, rest, re.DOTALL)
        if not options_found:
            continue
        first_opt_pos = re.search(r'[A-D]\)', rest)
        q_text = rest[:first_opt_pos.start()].strip() if first_opt_pos else rest
        opts = {letter: val.strip() for letter, val in options_found}
        questions.append({
            'id': qnum,
            'section': 'General',
            'passage': None,
            'question': q_text,
            'options': [opts.get('A',''), opts.get('B',''), opts.get('C',''), opts.get('D','')],
            'answer': ['A','B','C','D'].index(answers.get(qnum,'A')) if answers.get(qnum,'A') in ['A','B','C','D'] else 0,
            'image': None,
        })
    return questions

def build_analytics(filter_grade=None, filter_section=None, filter_subject=None):
    students    = User.query.filter_by(role='student').all()
    all_results = TestResult.query.all()

    # Apply filters
    if filter_grade:
        all_results = [r for r in all_results if r.student.grade == filter_grade]
        students    = [s for s in students if s.grade == filter_grade]
    if filter_section:
        all_results = [r for r in all_results if r.student.section == filter_section]
        students    = [s for s in students if s.section == filter_section]
    if filter_subject:
        all_results = [r for r in all_results if r.test.subject == filter_subject]

    overall_avg = safe_avg([r.percent for r in all_results])
    above80     = sum(1 for r in all_results if r.percent >= 80)
    below60     = sum(1 for r in all_results if r.percent < 60)

    grade_data = {}
    for g in GRADES:
        rs = [r for r in all_results if r.student.grade == g]
        grade_data[g] = {
            'avg': safe_avg([r.percent for r in rs]),
            'count': len(rs),
            'students': len([s for s in students if s.grade == g]),
        }

    subject_data = {}
    for sub in SUBJECTS:
        rs = [r for r in all_results if r.test.subject == sub]
        subject_data[sub] = {'avg': safe_avg([r.percent for r in rs]), 'count': len(rs)}

    grade_subject = {}
    for g in GRADES:
        grade_subject[g] = {}
        for sub in SUBJECTS:
            rs = [r for r in all_results if r.student.grade == g and r.test.subject == sub]
            grade_subject[g][sub] = safe_avg([r.percent for r in rs])

    # Strand-wise breakdown PER SUBJECT
    subject_strands = {}
    for sub in SUBJECTS:
        sub_results = [r for r in all_results if r.test.subject == sub]
        strand_data = {}
        for r in sub_results:
            try:
                secs = json.loads(r.section_scores or '{}')
                for sec, v in secs.items():
                    if v['total'] > 0:
                        pct = round(v['correct']/v['total']*100, 1)
                        strand_data.setdefault(sec, []).append(pct)
            except Exception:
                pass
        subject_strands[sub] = {sec: safe_avg(vals) for sec, vals in strand_data.items()}

    # Overall section avgs
    section_data = {}
    for r in all_results:
        try:
            secs = json.loads(r.section_scores or '{}')
            for sec, v in secs.items():
                if v['total'] > 0:
                    pct = round(v['correct']/v['total']*100, 1)
                    section_data.setdefault(sec, []).append(pct)
        except Exception:
            pass
    section_avgs = {sec: safe_avg(vals) for sec, vals in section_data.items()}

    student_rows = []
    for s in students:
        rs = [r for r in all_results if r.student_id == s.id]
        sub_avgs = {sub: safe_avg([r.percent for r in rs if r.test.subject == sub]) for sub in SUBJECTS}
        student_rows.append({
            'id': s.id, 'name': s.name, 'grade': s.grade, 'section': s.section,
            'tests_taken': len(rs),
            'overall_avg': safe_avg([r.percent for r in rs]),
            'sub_avgs': sub_avgs,
        })
    student_rows.sort(key=lambda x: -x['overall_avg'])

    return dict(
        overall_avg=overall_avg, above80=above80, below60=below60,
        total_results=len(all_results), total_students=len(students),
        grade_data=grade_data, subject_data=subject_data,
        grade_subject=grade_subject, section_avgs=section_avgs,
        subject_strands=subject_strands,
        student_rows=student_rows, subjects=SUBJECTS, grades=GRADES,
    )

def seed_db():
    try:
        if User.query.first():
            return
    except Exception:
        return
    db.session.add(User(name='Bushra Khan', username='Organizer',
        password=generate_password_hash('bk*123', method='pbkdf2:sha256:10000'),
        role='Resource_Manager'))
    for name, uname in [('Mrs. Sharma','teacher1'),('Mr. Verma','teacher2')]:
        db.session.add(User(name=name, username=uname,
            password=generate_password_hash('teacher123', method='pbkdf2:sha256:10000'),
            role='teacher'))
    sample = [
        ('Aarav Sharma','aarav','Grade 3','A'),('Priya Mehta','priya','Grade 3','A'),
        ('Rohan Gupta','rohan','Grade 4','B'),('Sneha Patel','sneha','Grade 4','A'),
    ]
    for name, uname, grade, sec in sample:
        db.session.add(User(name=name, username=uname,
            password=generate_password_hash('student123', method='pbkdf2:sha256:10000'),
            role='student', grade=grade, section=sec))
    db.session.commit()
    print("✅ Database seeded — Organizer / bk*123")

# ── HEALTH CHECK ──────────────────────────────────────────────────────────────

@app.route('/health')
def health():
    return 'OK', 200

# ── AUTH ──────────────────────────────────────────────────────────────────────

@app.route('/', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','').strip()
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            session['role']    = user.role
            session['name']    = user.name
            role = user.role
            if role == 'Resource_Manager':
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for(f"{role}_dashboard"))
        flash('Invalid username or password.', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ── ADMIN ─────────────────────────────────────────────────────────────────────

@app.route('/admin')
@login_required('Resource_Manager')
def admin_dashboard():
    students  = User.query.filter_by(role='student').all()
    tests     = MockTest.query.all()
    results   = TestResult.query.all()
    avg_score = safe_avg([r.percent for r in results])
    recent    = sorted(results, key=lambda r: r.taken_at, reverse=True)[:8]
    return render_template('admin/dashboard.html',
        students=students, tests=tests, results=results,
        avg_score=avg_score, recent=recent, subjects=SUBJECTS, grades=GRADES)

@app.route('/admin/students', methods=['GET','POST'])
@login_required('Resource_Manager')
def admin_students():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            if User.query.filter_by(username=request.form['username']).first():
                flash('Username already exists.', 'error')
            else:
                db.session.add(User(
                    name=request.form['name'], username=request.form['username'],
                    password=generate_password_hash(request.form['password'], method='pbkdf2:sha256:10000'),
                    role='student', grade=request.form['grade'], section=request.form.get('section','A')))
                db.session.commit()
                flash(f"Student {request.form['name']} added successfully.", 'success')
        elif action == 'delete':
            user = db.session.get(User, int(request.form['user_id']))
            if user:
                db.session.delete(user); db.session.commit()
                flash('Student removed.', 'success')
        elif action == 'edit':
            user = db.session.get(User, int(request.form['user_id']))
            if user:
                user.name = request.form['name']
                user.grade = request.form['grade']
                user.section = request.form.get('section','A')
                if request.form.get('password'):
                    user.password = generate_password_hash(request.form['password'], method='pbkdf2:sha256:10000')
                db.session.commit()
                flash('Student updated.', 'success')
    students = User.query.filter_by(role='student').order_by(User.grade, User.name).all()
    return render_template('admin/students.html', students=students, grades=GRADES)

@app.route('/admin/students/upload', methods=['GET','POST'])
@login_required('Resource_Manager')
def upload_students():
    preview = []
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'preview':
            file = request.files.get('csv_file')
            if not file or not file.filename.lower().endswith('.csv'):
                flash('Please upload a valid .csv file.', 'error')
                return redirect(url_for('upload_students'))
            stream = io.StringIO(file.stream.read().decode('utf-8-sig'))
            reader = csv.DictReader(stream)
            for row in reader:
                name    = row.get('name','').strip()
                grade   = row.get('grade','').strip()
                section = row.get('section','A').strip()
                if not name or not grade:
                    continue
                username = row.get('username','').strip() or generate_username(name, grade)
                password = row.get('password','').strip() or generate_password(name, grade)
                if grade.strip().isdigit():
                    grade = f"Grade {grade.strip()}"
                preview.append({'name': name, 'grade': grade, 'section': section,
                                'username': username, 'password': password})
            return render_template('admin/upload_students.html', preview=preview, grades=GRADES)
        elif action == 'confirm':
            names     = request.form.getlist('name')
            usernames = request.form.getlist('username')
            passwords = request.form.getlist('password')
            grades    = request.form.getlist('grade')
            sections  = request.form.getlist('section')
            added = 0; skipped = 0
            try:
                BATCH = 20
                for i in range(len(names)):
                    if not names[i] or not usernames[i] or not passwords[i]:
                        continue
                    if User.query.filter_by(username=usernames[i]).first():
                        skipped += 1; continue
                    hashed = generate_password_hash(passwords[i], method='pbkdf2:sha256:10000')
                    db.session.add(User(name=names[i], username=usernames[i], password=hashed,
                        role='student', grade=grades[i], section=sections[i]))
                    added += 1
                    if added % BATCH == 0:
                        db.session.commit()
                db.session.commit()
                msg = f'✅ {added} students added successfully!'
                if skipped:
                    msg += f' ({skipped} skipped — username already exists)'
                flash(msg, 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'❌ Error: {str(e)}', 'error')
            return redirect(url_for('admin_students'))
    return render_template('admin/upload_students.html', preview=preview, grades=GRADES)

@app.route('/admin/students/download-template')
@login_required('Resource_Manager')
def download_csv_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['name','username','password','grade','section'])
    writer.writerow(['Ahmed Khan','ahmed_01','ahm@01','Grade 3','A'])
    writer.writerow(['Sara Ali','sara_02','sar@02','Grade 4','B'])
    output.seek(0)
    return Response(output.getvalue(), mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=students_template.csv'})

@app.route('/admin/students/download-credentials')
@login_required('Resource_Manager')
def download_credentials():
    students = User.query.filter_by(role='student').order_by(User.grade, User.name).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Name','Grade','Section','Username'])
    for s in students:
        writer.writerow([s.name, s.grade, s.section, s.username])
    output.seek(0)
    return Response(output.getvalue(), mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=student_credentials.csv'})

@app.route('/admin/teachers', methods=['GET','POST'])
@login_required('Resource_Manager')
def admin_teachers():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            if User.query.filter_by(username=request.form['username']).first():
                flash('Username already exists.', 'error')
            else:
                db.session.add(User(name=request.form['name'], username=request.form['username'],
                    password=generate_password_hash(request.form['password'], method='pbkdf2:sha256:10000'),
                    role='teacher'))
                db.session.commit()
                flash('Teacher added.', 'success')
        elif action == 'delete':
            user = db.session.get(User, int(request.form['user_id']))
            if user: db.session.delete(user); db.session.commit()
            flash('Teacher removed.', 'success')
    teachers = User.query.filter_by(role='teacher').all()
    return render_template('admin/teachers.html', teachers=teachers)

@app.route('/admin/tests', methods=['GET','POST'])
@login_required('Resource_Manager')
def admin_tests():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            grade = request.form['grade']
            if grade == 'All Grades':
                for g in GRADES:
                    db.session.add(MockTest(
                        name=request.form['name'], subject=request.form['subject'],
                        grade=g, difficulty=request.form['difficulty'],
                        duration=int(request.form['duration']), status=request.form['status']))
                db.session.commit()
                flash('Test created for all grades.', 'success')
            else:
                db.session.add(MockTest(
                    name=request.form['name'], subject=request.form['subject'],
                    grade=grade, difficulty=request.form['difficulty'],
                    duration=int(request.form['duration']), status=request.form['status']))
                db.session.commit()
                flash('Test created.', 'success')
        elif action == 'edit':
            test_id = request.form.get('test_id')
            if test_id:
                t = db.session.get(MockTest, int(test_id))
                if t:
                    t.name       = request.form.get('name', t.name)
                    t.subject    = request.form.get('subject', t.subject)
                    t.grade      = request.form.get('grade', t.grade)
                    t.difficulty = request.form.get('difficulty', t.difficulty)
                    t.duration   = int(request.form.get('duration', t.duration))
                    t.status     = request.form.get('status', t.status)
                    db.session.commit()
                    flash('Test updated.', 'success')
        elif action == 'delete':
            test_id = request.form.get('test_id')
            if test_id:
                t = db.session.get(MockTest, int(test_id))
                if t:
                    db.session.delete(t); db.session.commit()
                    flash('Test deleted.', 'success')
        elif action == 'toggle':
            test_id = request.form.get('test_id')
            if test_id:
                t = db.session.get(MockTest, int(test_id))
                if t:
                    t.status = 'active' if t.status == 'draft' else 'draft'
                    db.session.commit()
    tests = MockTest.query.order_by(MockTest.created.desc()).all()
    all_grades = ['All Grades'] + GRADES
    return render_template('admin/tests.html', tests=tests, subjects=SUBJECTS, grades=GRADES, all_grades=all_grades)

@app.route('/admin/tests/<int:test_id>/questions', methods=['GET','POST'])
@login_required('Resource_Manager')
def admin_questions(test_id):
    test = db.session.get(MockTest, test_id)
    if not test: flash('Test not found.','error'); return redirect(url_for('admin_tests'))
    subject_sections = SECTIONS_BY_SUBJECT.get(test.subject, ['General'])
    if request.method == 'POST':
        qs = json.loads(test.questions or '[]')
        image_data = None
        img_file = request.files.get('question_image')
        if img_file and img_file.filename:
            ext = img_file.filename.rsplit('.', 1)[-1].lower()
            mime = {'jpg':'image/jpeg','jpeg':'image/jpeg','png':'image/png','gif':'image/gif','webp':'image/webp'}.get(ext,'image/png')
            raw = img_file.read()
            if len(raw) < 2 * 1024 * 1024:
                image_data = f"data:{mime};base64,{base64.b64encode(raw).decode()}"
            else:
                flash('Image too large — max 2MB.', 'error')
        qs.append({
            'id': max((q['id'] for q in qs), default=0) + 1,
            'section': request.form['section'],
            'passage': request.form.get('passage') or None,
            'question': request.form['question'],
            'options': [request.form.get(f'opt{i}','') for i in range(4)],
            'answer': int(request.form['answer']),
            'image': image_data,
        })
        test.questions = json.dumps(qs)
        db.session.commit()
        flash('Question added.', 'success')
    questions = json.loads(test.questions or '[]')
    return render_template('admin/questions.html', test=test, questions=questions, subject_sections=subject_sections)

@app.route('/admin/tests/<int:test_id>/upload-pdf', methods=['POST'])
@login_required('Resource_Manager')
def upload_pdf_questions(test_id):
    test = db.session.get(MockTest, test_id)
    if not test:
        flash('Test not found.', 'error')
        return redirect(url_for('admin_tests'))
    file = request.files.get('pdf_file')
    if not file or not file.filename.endswith('.pdf'):
        flash('Please upload a valid PDF file.', 'error')
        return redirect(url_for('admin_questions', test_id=test_id))
    try:
        parsed = parse_pdf_questions(file.stream)
        if not parsed:
            flash('No questions found in PDF.', 'error')
            return redirect(url_for('admin_questions', test_id=test_id))
        existing = json.loads(test.questions or '[]')
        max_id = max((q['id'] for q in existing), default=0)
        for q in parsed:
            q['id'] = max_id + q['id']
            existing.append(q)
        test.questions = json.dumps(existing)
        db.session.commit()
        flash(f'✅ {len(parsed)} questions imported from PDF!', 'success')
    except Exception as e:
        flash(f'❌ Error reading PDF: {str(e)}', 'error')
    return redirect(url_for('admin_questions', test_id=test_id))

@app.route('/admin/questions/edit/<int:test_id>/<int:q_id>', methods=['POST'])
@login_required('Resource_Manager')
def edit_question(test_id, q_id):
    test = db.session.get(MockTest, test_id)
    if not test:
        flash('Test not found.', 'error')
        return redirect(url_for('admin_tests'))
    qs = json.loads(test.questions or '[]')
    for q in qs:
        if q['id'] == q_id:
            q['question'] = request.form.get('question', q['question'])
            q['section']  = request.form.get('section', q.get('section','General'))
            q['passage']  = request.form.get('passage') or None
            q['options']  = [request.form.get(f'opt{i}', q['options'][i] if i < len(q['options']) else '') for i in range(4)]
            q['answer']   = int(request.form.get('answer', q['answer']))
            img_file = request.files.get('question_image')
            if img_file and img_file.filename:
                ext = img_file.filename.rsplit('.', 1)[-1].lower()
                mime = {'jpg':'image/jpeg','jpeg':'image/jpeg','png':'image/png','gif':'image/gif','webp':'image/webp'}.get(ext,'image/png')
                raw = img_file.read()
                if len(raw) < 2 * 1024 * 1024:
                    q['image'] = f"data:{mime};base64,{base64.b64encode(raw).decode()}"
                else:
                    flash('Image too large — max 2MB.', 'error')
            if request.form.get('remove_image'):
                q['image'] = None
            break
    test.questions = json.dumps(qs)
    db.session.commit()
    flash('Question updated.', 'success')
    return redirect(url_for('admin_questions', test_id=test_id))

@app.route('/admin/questions/delete/<int:test_id>/<int:q_id>', methods=['POST'])
@login_required('Resource_Manager')
def delete_question(test_id, q_id):
    test = db.session.get(MockTest, test_id)
    if test:
        qs = [q for q in json.loads(test.questions or '[]') if q['id'] != q_id]
        test.questions = json.dumps(qs)
        db.session.commit()
        flash('Question deleted.', 'success')
    return redirect(url_for('admin_questions', test_id=test_id))

@app.route('/admin/analytics')
@login_required('Resource_Manager')
def admin_analytics():
    filter_grade   = request.args.get('grade', '')
    filter_section = request.args.get('section', '')
    filter_subject = request.args.get('subject', '')
    data = build_analytics(
        filter_grade   = filter_grade   or None,
        filter_section = filter_section or None,
        filter_subject = filter_subject or None,
    )
    data['filter_grade']   = filter_grade
    data['filter_section'] = filter_section
    data['filter_subject'] = filter_subject
    data['sections']       = ['A','B','C','D']
    return render_template('admin/analytics.html', **data)

@app.route('/admin/download/students')
@login_required('Resource_Manager')
def download_students_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('openpyxl not installed.', 'error')
        return redirect(url_for('admin_students'))
    students = User.query.filter_by(role='student').order_by(User.grade, User.section, User.name).all()
    wb = Workbook(); ws = wb.active; ws.title = "Students"
    headers = ["#","Name","Username","Grade","Section","Tests Taken","Average %","Joined"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1a3c6e")
        c.alignment = Alignment(horizontal="center")
    for idx, s in enumerate(students, 1):
        results = s.results
        avg = round(sum(r.percent for r in results)/len(results), 1) if results else 0
        row = [idx, s.name, s.username, s.grade, s.section or '-', len(results), avg, s.created.strftime('%d %b %Y')]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=idx+1, column=c, value=val)
            cell.alignment = Alignment(horizontal="center")
            if idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EBF1F9")
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(len(str(c.value or '')) for c in col)+4, 40)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(buf.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=EPS_Students.xlsx'})

@app.route('/admin/download/results')
@login_required('Resource_Manager')
def download_results_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('openpyxl not installed.', 'error')
        return redirect(url_for('admin_analytics'))
    results = TestResult.query.order_by(TestResult.taken_at.desc()).all()
    wb = Workbook(); ws = wb.active; ws.title = "All Results"
    headers = ["#","Student Name","Username","Grade","Section","Test","Subject","Score","Total","%","Time","Date"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1a3c6e")
        c.alignment = Alignment(horizontal="center")
    for idx, r in enumerate(results, 1):
        mins = r.time_taken // 60; secs = r.time_taken % 60
        row = [idx, r.student.name, r.student.username, r.student.grade, r.student.section or '-',
               r.test.name, r.test.subject, r.score, r.total, r.percent,
               f"{mins}m {secs}s", r.taken_at.strftime('%d %b %Y')]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=idx+1, column=c, value=val)
            cell.alignment = Alignment(horizontal="center")
            if idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EBF1F9")
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(len(str(c.value or '')) for c in col)+4, 40)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(buf.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=EPS_Results.xlsx'})

# ── RESULTS TEMPLATE DOWNLOAD ────────────────────────────────────────────────

@app.route('/admin/download/results-template')
@login_required('Resource_Manager')
def download_results_template():
    tests = MockTest.query.all()
    # Get max questions across all tests
    max_q = max((len(json.loads(t.questions or '[]')) for t in tests), default=8)
    output = io.StringIO()
    writer = csv.writer(output)
    headers = ['username'] + [f'q{i+1}' for i in range(max_q)]
    writer.writerow(headers)
    # Sample rows
    students = User.query.filter_by(role='student').limit(3).all()
    for s in students:
        writer.writerow([s.username] + ['' for _ in range(max_q)])
    output.seek(0)
    return Response(output.getvalue(), mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=results_template.csv'})


# ── IB HOLISTIC ROUTES ───────────────────────────────────────────────────────

# ── Teacher: ATL Rating ───────────────────────────────────────────────────────
# ── IB HOLISTIC ROUTES ────────────────────────────────────────────────────────

@app.route('/ib')
@login_required('Resource_Manager')
def ib_dashboard():
    students = User.query.filter_by(role='student').all()
    total_atl = ATLRating.query.count()
    total_lp  = LearnerProfileRating.query.count()
    # Average LP ratings per attribute
    lp_avgs = {}
    for attr, _ in LEARNER_PROFILE:
        ratings = LearnerProfileRating.query.filter_by(attribute=attr).all()
        lp_avgs[attr] = round(sum(r.rating for r in ratings)/len(ratings),1) if ratings else 0
    return render_template('ib/dashboard.html',
        students=students, total_atl=total_atl, total_lp=total_lp,
        lp_avgs=lp_avgs, learner_profile=LEARNER_PROFILE,
        rating_scale=RATING_SCALE, rating_colors=RATING_COLORS,
        atl_skills=ATL_SKILLS, terms=TERMS)


@app.route('/ib/atl', methods=['GET','POST'])
@login_required('Resource_Manager')
def ib_atl():
    students = User.query.filter_by(role='student').order_by(User.grade, User.name).all()
    if request.method == 'POST':
        student_id = int(request.form.get('student_id'))
        term       = request.form.get('term')
        skill      = request.form.get('skill')
        rater_type = 'teacher'
        student    = db.session.get(User, student_id)
        grade      = student.grade if student else 'Grade 3'
        descriptors = ATL_SKILLS.get(skill, {}).get(grade, [])
        saved = 0
        for i, desc in enumerate(descriptors):
            rating_val = request.form.get(f'rating_{i}')
            if rating_val:
                existing = ATLRating.query.filter_by(
                    student_id=student_id, term=term,
                    skill=skill, descriptor=desc, rater_type=rater_type
                ).first()
                if existing:
                    existing.rating = int(rating_val)
                else:
                    db.session.add(ATLRating(
                        student_id=student_id, teacher_id=session['user_id'],
                        term=term, skill=skill, descriptor=desc,
                        rating=int(rating_val), rater_type=rater_type
                    ))
                saved += 1
        db.session.commit()
        flash(f'ATL ratings saved for {student.name} — {skill} ({term})', 'success')
        return redirect(url_for('ib_atl'))

    selected_student = request.args.get('student_id', type=int)
    selected_term    = request.args.get('term', 'Term 1')

    # Get existing ratings for selected student
    existing_ratings = {}
    if selected_student:
        ratings = ATLRating.query.filter_by(
            student_id=selected_student, term=selected_term, rater_type='teacher'
        ).all()
        for r in ratings:
            existing_ratings[(r.skill, r.descriptor)] = r.rating

    return render_template('ib/atl.html',
        students=students, terms=TERMS,
        atl_skills=ATL_SKILLS, rating_scale=RATING_SCALE,
        selected_student=selected_student, selected_term=selected_term,
        existing_ratings=existing_ratings)


@app.route('/ib/learner-profile', methods=['GET','POST'])
@login_required('Resource_Manager')
def ib_learner_profile():
    students = User.query.filter_by(role='student').order_by(User.grade, User.name).all()
    if request.method == 'POST':
        student_id = int(request.form.get('student_id'))
        term       = request.form.get('term')
        student    = db.session.get(User, student_id)
        for attr, _ in LEARNER_PROFILE:
            t_rating = request.form.get(f'teacher_{attr}')
            evidence = request.form.get(f'evidence_{attr}', '')
            if t_rating:
                existing = LearnerProfileRating.query.filter_by(
                    student_id=student_id, term=term,
                    attribute=attr, rater_type='teacher'
                ).first()
                if existing:
                    existing.rating = int(t_rating)
                    existing.evidence = evidence
                else:
                    db.session.add(LearnerProfileRating(
                        student_id=student_id, teacher_id=session['user_id'],
                        term=term, attribute=attr,
                        rating=int(t_rating), rater_type='teacher', evidence=evidence
                    ))
        db.session.commit()
        flash(f'Learner Profile saved for {student.name} ({term})', 'success')
        return redirect(url_for('ib_learner_profile'))

    selected_student = request.args.get('student_id', type=int)
    selected_term    = request.args.get('term', 'Term 1')
    existing_ratings = {}
    existing_evidence = {}
    if selected_student:
        ratings = LearnerProfileRating.query.filter_by(
            student_id=selected_student, term=selected_term, rater_type='teacher'
        ).all()
        for r in ratings:
            existing_ratings[r.attribute]  = r.rating
            existing_evidence[r.attribute] = r.evidence or ''

    return render_template('ib/learner_profile.html',
        students=students, terms=TERMS,
        learner_profile=LEARNER_PROFILE, rating_scale=RATING_SCALE,
        rating_colors=RATING_COLORS,
        selected_student=selected_student, selected_term=selected_term,
        existing_ratings=existing_ratings, existing_evidence=existing_evidence)


@app.route('/ib/report/<int:student_id>')
@login_required('Resource_Manager')
def ib_student_report(student_id):
    student = db.session.get(User, student_id)
    if not student:
        flash('Student not found.', 'error')
        return redirect(url_for('ib_dashboard'))
    report = {}
    for term in TERMS:
        atl_data = {}
        for skill in ATL_SKILLS:
            ratings = ATLRating.query.filter_by(
                student_id=student_id, term=term, skill=skill, rater_type='teacher'
            ).all()
            if ratings:
                atl_data[skill] = {
                    'avg': round(sum(r.rating for r in ratings)/len(ratings), 1),
                    'descriptors': [(r.descriptor, r.rating) for r in ratings]
                }
        lp_data = {}
        for attr, desc in LEARNER_PROFILE:
            t_r = LearnerProfileRating.query.filter_by(
                student_id=student_id, term=term, attribute=attr, rater_type='teacher'
            ).first()
            s_r = LearnerProfileRating.query.filter_by(
                student_id=student_id, term=term, attribute=attr, rater_type='student'
            ).first()
            if t_r or s_r:
                lp_data[attr] = {
                    'teacher': t_r.rating if t_r else None,
                    'student': s_r.rating if s_r else None,
                    'evidence': t_r.evidence if t_r else '',
                }
        report[term] = {'atl': atl_data, 'lp': lp_data}

    test_results = TestResult.query.filter_by(student_id=student_id).all()
    return render_template('ib/student_report.html',
        student=student, report=report, terms=TERMS,
        learner_profile=LEARNER_PROFILE, atl_skills=list(ATL_SKILLS.keys()),
        rating_scale=RATING_SCALE, rating_colors=RATING_COLORS,
        test_results=test_results)


# ── STUDENT IB SELF-RATING ─────────────────────────────────────────────────────

@app.route('/student/ib', methods=['GET','POST'])
@login_required('student')
def student_ib():
    student  = db.session.get(User, session['user_id'])
    selected_term = request.args.get('term', 'Term 1')

    if request.method == 'POST':
        term = request.form.get('term')
        for attr, _ in LEARNER_PROFILE:
            s_rating    = request.form.get(f'self_{attr}')
            reflection  = request.form.get(f'reflection_{attr}', '')
            if s_rating:
                existing = LearnerProfileRating.query.filter_by(
                    student_id=student.id, term=term,
                    attribute=attr, rater_type='student'
                ).first()
                if existing:
                    existing.rating = int(s_rating)
                else:
                    db.session.add(LearnerProfileRating(
                        student_id=student.id, term=term,
                        attribute=attr, rating=int(s_rating), rater_type='student'
                    ))
            if reflection:
                existing_r = StudentReflection.query.filter_by(
                    student_id=student.id, term=term, attribute=attr
                ).first()
                if existing_r:
                    existing_r.reflection = reflection
                else:
                    db.session.add(StudentReflection(
                        student_id=student.id, term=term,
                        attribute=attr, reflection=reflection
                    ))
        db.session.commit()
        flash('Your self-assessment has been saved!', 'success')
        return redirect(url_for('student_ib', term=term))

    # Load existing self-ratings and reflections
    self_ratings = {}
    reflections  = {}
    ratings = LearnerProfileRating.query.filter_by(
        student_id=student.id, term=selected_term, rater_type='student'
    ).all()
    for r in ratings:
        self_ratings[r.attribute] = r.rating

    teacher_ratings = {}
    t_ratings = LearnerProfileRating.query.filter_by(
        student_id=student.id, term=selected_term, rater_type='teacher'
    ).all()
    for r in t_ratings:
        teacher_ratings[r.attribute] = r.rating

    refs = StudentReflection.query.filter_by(
        student_id=student.id, term=selected_term
    ).all()
    for r in refs:
        reflections[r.attribute] = r.reflection

    # ATL ratings from teacher
    atl_ratings = {}
    atl = ATLRating.query.filter_by(
        student_id=student.id, term=selected_term, rater_type='teacher'
    ).all()
    for r in atl:
        atl_ratings.setdefault(r.skill, []).append({'desc': r.descriptor, 'rating': r.rating})

    return render_template('student/ib.html',
        student=student, terms=TERMS, selected_term=selected_term,
        learner_profile=LEARNER_PROFILE, rating_scale=RATING_SCALE,
        rating_colors=RATING_COLORS,
        self_ratings=self_ratings, teacher_ratings=teacher_ratings,
        reflections=reflections, atl_ratings=atl_ratings)



# ── TEACHER ───────────────────────────────────────────────────────────────────

@app.route('/teacher')
@login_required('teacher')
def teacher_dashboard():
    students  = User.query.filter_by(role='student').all()
    results   = TestResult.query.all()
    tests     = MockTest.query.filter_by(status='active').all()
    avg_score = safe_avg([r.percent for r in results])
    recent    = sorted(results, key=lambda r: r.taken_at, reverse=True)[:6]
    return render_template('teacher/dashboard.html',
        students=students, results=results, tests=tests,
        avg_score=avg_score, recent=recent)

@app.route('/teacher/students')
@login_required('teacher')
def teacher_students():
    students = User.query.filter_by(role='student').order_by(User.grade, User.name).all()
    student_data = []
    for s in students:
        rs = list(s.results)
        student_data.append({
            'student': s, 'tests_taken': len(rs),
            'avg': safe_avg([r.percent for r in rs]),
        })
    return render_template('teacher/students.html', student_data=student_data)

@app.route('/teacher/analytics')
@login_required('teacher')
def teacher_analytics():
    data = build_analytics()
    return render_template('teacher/analytics.html', **data)

# ── STUDENT ───────────────────────────────────────────────────────────────────

@app.route('/student')
@login_required('student')
def student_dashboard():
    student = db.session.get(User, session['user_id'])
    seen_tests = set()
    unique_results = []
    for r in sorted(student.results, key=lambda r: r.taken_at):
        if r.test_id not in seen_tests:
            seen_tests.add(r.test_id)
            unique_results.append(r)
    results = sorted(unique_results, key=lambda r: r.taken_at, reverse=True)
    tests   = MockTest.query.filter(
        MockTest.status=='active',
        db.or_(MockTest.grade==student.grade, MockTest.grade=='All Grades')
    ).all()
    completed_test_ids = [r.test_id for r in results]
    avg_sc = safe_avg([r.percent for r in results])
    diagnostic_insights = dt_student_insights(dt_student_series(student.id))
    return render_template('student/dashboard.html',
        student=student, results=results, tests=tests,
        completed_test_ids=completed_test_ids, avg_sc=avg_sc, subjects=SUBJECTS,
        diagnostic_insights=diagnostic_insights)

@app.route('/student/test/<int:test_id>')
@login_required('student')
def student_test(test_id):
    test = db.session.get(MockTest, test_id)
    if not test: return redirect(url_for('student_dashboard'))
    existing = TestResult.query.filter_by(student_id=session['user_id'], test_id=test_id).first()
    if existing:
        flash('You have already completed this test.', 'error')
        return redirect(url_for('student_dashboard'))
    questions = json.loads(test.questions or '[]')
    return render_template('student/test.html', test=test, questions=questions)

@app.route('/student/submit/<int:test_id>', methods=['POST'])
@login_required('student')
def submit_test(test_id):
    test = db.session.get(MockTest, test_id)
    if not test: return jsonify({'error':'not found'}), 404
    existing = TestResult.query.filter_by(student_id=session['user_id'], test_id=test_id).first()
    if existing:
        return jsonify({'score':existing.score,'total':existing.total,'percent':existing.percent,'section_scores':json.loads(existing.section_scores or '{}')})
    questions  = json.loads(test.questions or '[]')
    data       = request.get_json() or {}
    answers    = data.get('answers', {})
    time_taken = data.get('time_taken', 0)
    score = 0
    section_scores = {}
    for q in questions:
        sec = q.get('section', 'General')
        section_scores.setdefault(sec, {'correct':0,'total':0})
        section_scores[sec]['total'] += 1
        if str(q['id']) in answers and answers[str(q['id'])] == q['answer']:
            score += 1
            section_scores[sec]['correct'] += 1
    total   = len(questions)
    percent = round(score/total*100, 1) if total else 0
    db.session.add(TestResult(
        student_id=session['user_id'], test_id=test_id,
        score=score, total=total, percent=percent,
        answers=json.dumps(answers), section_scores=json.dumps(section_scores),
        time_taken=time_taken))
    db.session.commit()
    return jsonify({'score':score,'total':total,'percent':percent,'section_scores':section_scores})

@app.route('/student/scores')
@login_required('student')
def student_scores():
    student = db.session.get(User, session['user_id'])
    seen_tests = set()
    unique_results = []
    for r in sorted(student.results, key=lambda r: r.taken_at):
        if r.test_id not in seen_tests:
            seen_tests.add(r.test_id)
            unique_results.append(r)
    results = sorted(unique_results, key=lambda r: r.taken_at, reverse=True)
    return render_template('student/scores.html', student=student, results=results)

@app.route('/student/review/<int:result_id>')
@login_required('student')
def student_review(result_id):
    result  = db.session.get(TestResult, result_id)
    if not result:
        flash('Result not found.', 'error')
        return redirect(url_for('student_scores'))
    student = db.session.get(User, session['user_id'])
    if result.student_id != student.id:
        flash('Access denied.', 'error')
        return redirect(url_for('student_scores'))
    test      = result.test
    questions = json.loads(test.questions or '[]')
    answers   = json.loads(result.answers or '{}')
    review = []
    for q in questions:
        qid       = str(q['id'])
        given_idx = answers.get(qid)
        correct   = q.get('answer', 0)
        if given_idx is None:
            status = 'unattempted'
        elif int(given_idx) == correct:
            status = 'correct'
        else:
            status = 'wrong'
        review.append({'question':q,'given':int(given_idx) if given_idx is not None else None,'correct':correct,'status':status})
    return render_template('student/review.html', student=student, test=test, result=result, review=review)

# ── API ───────────────────────────────────────────────────────────────────────

@app.route('/api/analytics')
@login_required('Resource_Manager')
def api_analytics():
    results  = TestResult.query.all()
    by_grade = {}
    for r in results:
        g = r.student.grade or 'Unknown'
        by_grade.setdefault(g, []).append(r.percent)
    return jsonify({g: round(sum(v)/len(v),1) for g,v in by_grade.items()})

# ── IMPORT RESULTS ────────────────────────────────────────────────────────────

@app.route('/admin/import-results', methods=['GET','POST'])
@login_required('Resource_Manager')
def import_results():
    tests   = MockTest.query.all()
    preview = []
    errors  = []
    if request.method == 'POST':
        action  = request.form.get('action')
        test_id = int(request.form.get('test_id', 0))
        test    = db.session.get(MockTest, test_id)
        if not test:
            flash('Test not found.', 'error')
            return redirect(url_for('import_results'))
        questions = json.loads(test.questions or '[]')
        total_q   = len(questions)
        opt_map   = {'A':0,'B':1,'C':2,'D':3,'a':0,'b':1,'c':2,'d':3}
        if action == 'preview':
            file = request.files.get('csv_file')
            if not file or not file.filename.endswith('.csv'):
                flash('Please upload a valid .csv file.', 'error')
                return redirect(url_for('import_results'))
            stream = io.StringIO(file.stream.read().decode('utf-8-sig'))
            reader = csv.DictReader(stream)
            for row in reader:
                username = row.get('username','').strip()
                if not username: continue
                user = User.query.filter_by(username=username).first()
                if not user:
                    errors.append(f"'{username}' not found — skipped")
                    continue
                existing = TestResult.query.filter_by(student_id=user.id, test_id=test_id).first()
                if existing:
                    errors.append(f"{username} already has a result — skipped")
                    continue
                answers = {}
                score   = 0
                section_scores = {}
                for i, q in enumerate(questions):
                    col         = f'q{i+1}'
                    given_letter = row.get(col,'').strip().upper()
                    given_idx   = opt_map.get(given_letter)
                    correct_idx = q['answer']
                    sec         = q.get('section','General')
                    section_scores.setdefault(sec, {'correct':0,'total':0})
                    section_scores[sec]['total'] += 1
                    if given_idx is not None:
                        answers[str(q['id'])] = given_idx
                        if given_idx == correct_idx:
                            score += 1
                            section_scores[sec]['correct'] += 1
                percent = round(score/total_q*100, 1) if total_q else 0
                preview.append({'username':username,'name':user.name,'user_id':user.id,'score':score,'total':total_q,'percent':percent,'answers':json.dumps(answers),'section_scores':json.dumps(section_scores)})
            return render_template('admin/import_results.html', tests=tests, preview=preview, errors=errors, test_id=test_id, selected_test=test)
        elif action == 'confirm':
            userids=request.form.getlist('user_id'); scores=request.form.getlist('score')
            totals=request.form.getlist('total'); percents=request.form.getlist('percent')
            answers_l=request.form.getlist('answers'); secs_l=request.form.getlist('section_scores')
            added = 0
            for i in range(len(userids)):
                existing = TestResult.query.filter_by(student_id=int(userids[i]), test_id=test_id).first()
                if existing: continue
                db.session.add(TestResult(student_id=int(userids[i]),test_id=test_id,score=int(scores[i]),total=int(totals[i]),percent=float(percents[i]),answers=answers_l[i],section_scores=secs_l[i],time_taken=0))
                added += 1
            db.session.commit()
            flash(f"✅ {added} student results imported successfully!", 'success')
            return redirect(url_for('admin_analytics'))
    return render_template('admin/import_results.html', tests=tests, preview=[], errors=[], test_id=None, selected_test=None)

@app.route('/admin/download/test/<int:test_id>')
@login_required('Resource_Manager')
def download_test_excel(test_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    test    = db.session.get(MockTest, test_id)
    if not test:
        flash('Test not found.', 'error')
        return redirect(url_for('admin_analytics'))
    results   = TestResult.query.filter_by(test_id=test_id).order_by(TestResult.percent.desc()).all()
    questions = json.loads(test.questions or '[]')
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    headers = ["#","Name","Grade","Section","Score","Total","%","Time"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1a3c6e")
        c.alignment = Alignment(horizontal="center")
    for idx, r in enumerate(results, 1):
        mins = r.time_taken // 60; secs = r.time_taken % 60
        row = [idx, r.student.name, r.student.grade, r.student.section or '-', r.score, r.total, r.percent, f"{mins}m {secs}s"]
        for c, val in enumerate(row, 1):
            ws.cell(row=idx+1, column=c, value=val)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = test.name.replace(' ','_')
    return Response(buf.read(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=EPS_{fname}.xlsx'})

# ── MAIN ──────────────────────────────────────────────────────────────────────


# ═══════════════════════════════════════════════════════════════════════════
# DT (DIAGNOSTIC TEST) MODULE — PATCH FOR app.py
# ═══════════════════════════════════════════════════════════════════════════
#
# WHERE TO PASTE EACH SECTION:
#   1. "NEW IMPORTS"              -> top of app.py, with your other imports
#   2. "UPDATED login_required"   -> REPLACES your existing login_required()
#   3. "NEW CONSTANTS"            -> near your SUBJECTS/GRADES constants
#   4. "NEW MODELS"               -> in the MODELS section, after TestResult
#   5. "NEW HELPERS"              -> in the HELPERS section, after safe_avg()
#   6. "NEW ROUTES"               -> anywhere after the teacher/admin routes
#
# ALSO REQUIRED:
#   - Add `reportlab` to requirements.txt (see file 6)
#   - Add the sidebar nav snippet to base.html (see file 5)
#   - Drop templates/dt/entry.html, upload.html, graph.html into templates/dt/
#
# ═══════════════════════════════════════════════════════════════════════════


# ── 1. NEW IMPORTS ────────────────────────────────────────────────────────
# Add near the top of app.py with your other imports:
#
#   from reportlab.pdfgen import canvas as pdfcanvas
#   from reportlab.lib.pagesizes import A4
#   from reportlab.lib import colors as rl_colors
#   from reportlab.graphics.shapes import Drawing
#   from reportlab.graphics.charts.linecharts import HorizontalLineChart
#   from reportlab.graphics import renderPDF


# ── 2. UPDATED login_required — allows a role OR a tuple of roles ─────────
# REPLACE your existing login_required() with this version (backward compatible
# — passing a single string still works exactly as before):

def login_required(role=None):
    from functools import wraps
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            if role:
                allowed_roles = (
                    role
                    if isinstance(role, (list, tuple))
                    else (role,)
                )
                if session.get('role') not in allowed_roles:
                    flash('Access denied.', 'error')
                    return redirect(url_for('login'))
            return f(*args, **kwargs)
        return decorated
    return decorator
# ── DT CONSTANTS ─────────────────────────────────────────────────────────────
ACADEMIC_YEAR = '2025-26'
DT_NUMBERS = [
    1,
    2,
    3,
    4,
    5,
    6
]
DT_SUBJECTS = [
    'English',
    'Hindi',
    'Maths',
    'Science',
    'Urdu',
    'ICT'
]
DT_GRADES = [
    'Grade 1',
    'Grade 2',
    'Grade 3',
    'Grade 4',
    'Grade 5'
]
DT_SECTIONS = [
    'A',
    'B',
    'C',
    'D'
]
# ── DT MODELS ────────────────────────────────────────────────────────────────
class DiagnosticTest(db.Model):
    __tablename__ = 'diagnostic_test'
    id = db.Column(
        db.Integer,
        primary_key=True
    )
    dt_number = db.Column(
        db.Integer,
        nullable=False
    )
    subject = db.Column(
        db.String(50),
        nullable=False
    )
    grade = db.Column(
        db.String(20),
        nullable=False
    )
    section = db.Column(
        db.String(10),
        nullable=True
    )
    max_marks = db.Column(
        db.Float,
        default=25
    )
    academic_year = db.Column(
        db.String(20),
        default=ACADEMIC_YEAR
    )
    test_date = db.Column(
        db.Date,
        nullable=True
    )
    created_by = db.Column(
        db.Integer,
        db.ForeignKey('user.id'),
        nullable=True
    )
    created = db.Column(
        db.DateTime,
        default=datetime.utcnow
    )
    marks = db.relationship(
        'DTMark',
        backref='dt',
        lazy=True,
        cascade='all,delete-orphan'
    )
    __table_args__ = (
        db.UniqueConstraint(
            'dt_number',
            'subject',
            'grade',
            'section',
            'academic_year',
            name='uq_dt_slot'
        ),
    )
class DTMark(db.Model):
    __tablename__ = 'dt_mark'
    id = db.Column(
        db.Integer,
        primary_key=True
    )
    dt_id = db.Column(
        db.Integer,
        db.ForeignKey('diagnostic_test.id'),
        nullable=False
    )
    student_id = db.Column(
        db.Integer,
        db.ForeignKey('user.id'),
        nullable=False
    )
    marks_obtained = db.Column(
        db.Float,
        nullable=False
    )
    remarks = db.Column(
        db.Text,
        nullable=True
    )
    entered_by = db.Column(
        db.Integer,
        db.ForeignKey('user.id'),
        nullable=True
    )
    entered_at = db.Column(
        db.DateTime,
        default=datetime.utcnow
    )
    student = db.relationship(
        'User',
        foreign_keys=[student_id]
    )
    __table_args__ = (
        db.UniqueConstraint(
            'dt_id',
            'student_id',
            name='uq_dt_student'
        ),
    )
# ── DT HELPERS ───────────────────────────────────────────────────────────────
def dt_get_or_create(
    dt_number,
    subject,
    grade,
    section,
    academic_year=ACADEMIC_YEAR,
    max_marks=25,
    created_by=None
):
    dt = DiagnosticTest.query.filter_by(
        dt_number=dt_number,
        subject=subject,
        grade=grade,
        section=section,
        academic_year=academic_year
    ).first()
    if not dt:
        dt = DiagnosticTest(
            dt_number=dt_number,
            subject=subject,
            grade=grade,
            section=section,
            academic_year=academic_year,
            max_marks=max_marks,
            created_by=created_by
        )
        db.session.add(dt)
        db.session.commit()
    return dt
def dt_student_series(
    student_id,
    academic_year=ACADEMIC_YEAR
):
    """
    Returns:
    {
        'English': [
            {
                'dt': 1,
                'marks': 20,
                'max': 25,
                'pct': 80.0,
                'class_avg_pct': 72.5,
                'date': ...
            }
        ]
    }
    """
    student = db.session.get(User, student_id)
    if not student:
        return {
            subject: []
            for subject in DT_SUBJECTS
        }
    series = {
        subject: []
        for subject in DT_SUBJECTS
    }
    for subject in DT_SUBJECTS:
        for dt_number in DT_NUMBERS:
            dt = DiagnosticTest.query.filter_by(
                dt_number=dt_number,
                subject=subject,
                grade=student.grade,
                academic_year=academic_year
            ).filter(
                db.or_(
                    DiagnosticTest.section == None,
                    DiagnosticTest.section == student.section
                )
            ).first()
            if not dt:
                series[subject].append({
                    'dt': dt_number,
                    'marks': None,
                    'max': None,
                    'pct': None,
                    'class_avg_pct': None,
                    'date': None
                })
                continue
            mark = DTMark.query.filter_by(
                dt_id=dt.id,
                student_id=student_id
            ).first()
            class_marks = [
                item.marks_obtained
                for item in dt.marks
            ]
            class_avg_pct = (
                round(
                    sum(class_marks)
                    / len(class_marks)
                    / dt.max_marks
                    * 100,
                    1
                )
                if class_marks and dt.max_marks
                else None
            )
            if mark:
                pct = (
                    round(
                        mark.marks_obtained
                        / dt.max_marks
                        * 100,
                        1
                    )
                    if dt.max_marks
                    else 0
                )
                series[subject].append({
                    'dt': dt_number,
                    'marks': mark.marks_obtained,
                    'max': dt.max_marks,
                    'pct': pct,
                    'class_avg_pct': class_avg_pct,
                    'date': dt.test_date
                })
            else:
                series[subject].append({
                    'dt': dt_number,
                    'marks': None,
                    'max': dt.max_marks,
                    'pct': None,
                    'class_avg_pct': class_avg_pct,
                    'date': dt.test_date
                })
    return series
def dt_student_insights(series):
    """
    Summarises diagnostic progress for dashboards and reports.
    """
    insights = []
    for subject in DT_SUBJECTS:
        points = [
            point
            for point in series.get(subject, [])
            if point.get('pct') is not None
        ]
        values = [
            point['pct']
            for point in points
        ]
        average = (
            round(sum(values) / len(values), 1)
            if values
            else None
        )
        trend = (
            round(values[-1] - values[0], 1)
            if len(values) > 1
            else None
        )
        if average is not None and average >= 80:
            status = 'Strong'
        elif average is not None and average >= 60:
            status = 'Developing'
        elif average is not None:
            status = 'Needs focus'
        else:
            status = 'Not started'
        insights.append({
            'subject': subject,
            'average': average,
            'trend': trend,
            'completed': len(points),
            'latest': values[-1] if values else None,
            'status': status
        })
    return sorted(
        insights,
        key=lambda item: (
            item['average'] is None,
            item['average'] or 0
        )
    )
def _pdf_header(c, title, subtitle):
    width, height = A4
    c.setFillColorRGB(
        0.10,
        0.24,
        0.43
    )
    c.rect(
        0,
        height - 90,
        width,
        90,
        fill=1,
        stroke=0
    )
    c.setFillColorRGB(
        1,
        1,
        1
    )
    c.setFont(
        'Helvetica-Bold',
        18
    )
    c.drawString(
        40,
        height - 45,
        'Eastern Public School'
    )
    c.setFont(
        'Helvetica',
        11
    )
    c.drawString(
        40,
        height - 65,
        title
    )
    c.setFont(
        'Helvetica',
        9
    )
    c.drawString(
        40,
        height - 80,
        subtitle
    )
    c.setFillColorRGB(
        0,
        0,
        0
    )
    return height
def _dt_role_prefix():
    if session.get('role') == 'Resource_Manager':
        return 'admin'
    return 'teacher'
# ── DT MARK ENTRY ────────────────────────────────────────────────────────────
@app.route(
    '/teacher/dt',
    methods=['GET', 'POST'],
    endpoint='teacher_dt'
)
@app.route(
    '/admin/dt',
    methods=['GET', 'POST'],
    endpoint='admin_dt'
)
@login_required(
    ('teacher', 'Resource_Manager')
)
def dt_entry():
    if request.method == 'POST':
        grade = request.form['grade']
        section = request.form.get('section') or None
        subject = request.form['subject']
        dt_number = int(request.form['dt_number'])
        max_marks = float(
            request.form.get(
                'max_marks',
                25
            )
        )
        test_date = request.form.get(
            'test_date'
        ) or None
        dt = dt_get_or_create(
            dt_number=dt_number,
            subject=subject,
            grade=grade,
            section=section,
            academic_year=ACADEMIC_YEAR,
            max_marks=max_marks,
            created_by=session['user_id']
        )
        dt.max_marks = max_marks
        if test_date:
            dt.test_date = datetime.strptime(
                test_date,
                '%Y-%m-%d'
            ).date()
        db.session.commit()
        student_ids = request.form.getlist(
            'student_id'
        )
        saved = 0
        for student_id in student_ids:
            value = request.form.get(
                f'marks_{student_id}',
                ''
            ).strip()
            if value == '':
                continue
            try:
                marks_value = float(value)
            except ValueError:
                continue
            remark = request.form.get(
                f'remark_{student_id}',
                ''
            ).strip()
            existing = DTMark.query.filter_by(
                dt_id=dt.id,
                student_id=int(student_id)
            ).first()
            if existing:
                existing.marks_obtained = marks_value
                existing.remarks = remark
                existing.entered_by = session['user_id']
                existing.entered_at = datetime.utcnow()
            else:
                db.session.add(
                    DTMark(
                        dt_id=dt.id,
                        student_id=int(student_id),
                        marks_obtained=marks_value,
                        remarks=remark,
                        entered_by=session['user_id']
                    )
                )
            saved += 1
        db.session.commit()
        flash(
            f'✅ Marks saved for {saved} student(s) — '
            f'{subject} DT{dt_number}, '
            f'{grade}{(" " + section) if section else ""}',
            'success'
        )
        return redirect(
            url_for(
                f'{_dt_role_prefix()}_dt',
                grade=grade,
                section=section or '',
                subject=subject,
                dt_number=dt_number
            )
        )
    grade = request.args.get(
        'grade',
        GRADES[0]
    )
    section = request.args.get(
        'section',
        ''
    )
    subject = request.args.get(
        'subject',
        DT_SUBJECTS[0]
    )
    dt_number = int(
        request.args.get(
            'dt_number',
            1
        )
    )
    query = User.query.filter_by(
        role='student',
        grade=grade
    )
    if section:
        query = query.filter_by(
            section=section
        )
    students = query.order_by(
        User.name
    ).all()
    dt = DiagnosticTest.query.filter_by(
        dt_number=dt_number,
        subject=subject,
        grade=grade,
        section=section or None,
        academic_year=ACADEMIC_YEAR
    ).first()
    existing_marks = {}
    if dt:
        for mark in dt.marks:
            existing_marks[mark.student_id] = {
                'marks': mark.marks_obtained,
                'remarks': mark.remarks or ''
            }
    sections = sorted({
        student.section
        for student in User.query.filter_by(
            role='student',
            grade=grade
        ).all()
        if student.section
    })
    return render_template(
        'dt/entry.html',
        students=students,
        grades=GRADES,
        subjects=DT_SUBJECTS,
        dt_numbers=DT_NUMBERS,
        sections=sections,
        grade=grade,
        section=section,
        subject=subject,
        dt_number=dt_number,
        dt=dt,
        existing_marks=existing_marks,
        academic_year=ACADEMIC_YEAR,
        role_prefix=_dt_role_prefix()
    )
# ── DT CSV UPLOAD ────────────────────────────────────────────────────────────
@app.route(
    '/teacher/dt/upload',
    methods=['GET', 'POST'],
    endpoint='teacher_dt_upload'
)
@app.route(
    '/admin/dt/upload',
    methods=['GET', 'POST'],
    endpoint='admin_dt_upload'
)
@login_required(
    ('teacher', 'Resource_Manager')
)
def dt_upload():
    grade = request.values.get(
        'grade',
        GRADES[0]
    )
    section = request.values.get(
        'section',
        ''
    )
    subject = request.values.get(
        'subject',
        DT_SUBJECTS[0]
    )
    dt_number = int(
        request.values.get(
            'dt_number',
            1
        )
    )
    max_marks = float(
        request.values.get(
            'max_marks',
            25
        )
    )
    if request.method == 'POST':
        file = request.files.get(
            'csv_file'
        )
        if not file or not file.filename.lower().endswith('.csv'):
            flash(
                'Please upload a valid .csv file.',
                'error'
            )
            return redirect(
                url_for(
                    f'{_dt_role_prefix()}_dt_upload',
                    grade=grade,
                    section=section,
                    subject=subject,
                    dt_number=dt_number,
                    max_marks=max_marks
                )
            )
        dt = dt_get_or_create(
            dt_number=dt_number,
            subject=subject,
            grade=grade,
            section=section or None,
            academic_year=ACADEMIC_YEAR,
            max_marks=max_marks,
            created_by=session['user_id']
        )
        stream = io.StringIO(
            file.stream.read().decode(
                'utf-8-sig'
            )
        )
        reader = csv.DictReader(
            stream
        )
        added = 0
        skipped = 0
        for row in reader:
            username = row.get(
                'username',
                ''
            ).strip()
            marks_raw = row.get(
                'marks',
                ''
            ).strip()
            if not username or marks_raw == '':
                continue
            student = User.query.filter_by(
                username=username,
                role='student'
            ).first()
            if not student:
                skipped += 1
                continue
            try:
                marks_value = float(
                    marks_raw
                )
                if marks_value < 0 or marks_value > max_marks:
                    skipped += 1
                    continue
            except ValueError:
                skipped += 1
                continue
            existing = DTMark.query.filter_by(
                dt_id=dt.id,
                student_id=student.id
            ).first()
            remarks = row.get(
                'remarks',
                ''
            ).strip()
            if existing:
                existing.marks_obtained = marks_value
                existing.remarks = remarks
                existing.entered_by = session['user_id']
                existing.entered_at = datetime.utcnow()
            else:
                db.session.add(
                    DTMark(
                        dt_id=dt.id,
                        student_id=student.id,
                        marks_obtained=marks_value,
                        remarks=remarks,
                        entered_by=session['user_id']
                    )
                )
            added += 1
        db.session.commit()
        flash(
            f'✅ {added} marks uploaded '
            f'({skipped} skipped — check usernames/marks)',
            'success'
        )
        return redirect(
            url_for(
                f'{_dt_role_prefix()}_dt',
                grade=grade,
                section=section,
                subject=subject,
                dt_number=dt_number
            )
        )
    return render_template(
        'dt/upload.html',
        grade=grade,
        section=section,
        subject=subject,
        dt_number=dt_number,
        max_marks=max_marks,
        grades=GRADES,
        subjects=DT_SUBJECTS,
        dt_numbers=DT_NUMBERS,
        role_prefix=_dt_role_prefix()
    )
# ── DT CSV TEMPLATE ─────────────────────────────────────────────────────────
@app.route(
    '/teacher/dt/upload-template',
    endpoint='teacher_dt_upload_template'
)
@app.route(
    '/admin/dt/upload-template',
    endpoint='admin_dt_upload_template'
)
@login_required(
    ('teacher', 'Resource_Manager')
)
def dt_upload_template():
    grade = request.args.get(
        'grade',
        ''
    )
    section = request.args.get(
        'section',
        ''
    )
    output = io.StringIO()
    writer = csv.writer(
        output
    )
    writer.writerow([
        'username',
        'marks',
        'remarks'
    ])
    query = User.query.filter_by(
        role='student'
    )
    if grade:
        query = query.filter_by(
            grade=grade
        )
    if section:
        query = query.filter_by(
            section=section
        )
    for student in query.order_by(
        User.name
    ).all():
        writer.writerow([
            student.username,
            '',
            ''
        ])
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition':
                'attachment; filename=DT_marks_template.csv'
        }
    )
# ── DT GRAPH ─────────────────────────────────────────────────────────────────
@app.route(
    '/teacher/dt/graph',
    endpoint='teacher_dt_graph'
)
@app.route(
    '/admin/dt/graph',
    endpoint='admin_dt_graph'
)
@login_required(
    ('teacher', 'Resource_Manager')
)
def dt_graph():
    student_id = request.args.get(
        'student_id',
        type=int
    )
    grade = request.args.get(
        'grade',
        GRADES[0]
    )
    section = request.args.get(
        'section',
        ''
    )
    query = User.query.filter_by(
        role='student',
        grade=grade
    )
    if section:
        query = query.filter_by(
            section=section
        )
    students = query.order_by(
        User.name
    ).all()
    sections = sorted({
        student.section
        for student in User.query.filter_by(
            role='student',
            grade=grade
        ).all()
        if student.section
    })
    series = None
    student = None
    if student_id:
        student = db.session.get(
            User,
            student_id
        )
        if student:
            series = dt_student_series(
                student_id
            )
    return render_template(
        'dt/graph.html',
        students=students,
        grades=GRADES,
        sections=sections,
        grade=grade,
        section=section,
        student=student,
        series=series,
        subjects=DT_SUBJECTS,
        dt_numbers=DT_NUMBERS,
        academic_year=ACADEMIC_YEAR,
        role_prefix=_dt_role_prefix()
    )
# ── SINGLE DT PDF ────────────────────────────────────────────────────────────
@app.route(
    '/teacher/dt/pdf/<int:student_id>/<int:dt_number>',
    endpoint='teacher_dt_pdf_single'
)
@app.route(
    '/admin/dt/pdf/<int:student_id>/<int:dt_number>',
    endpoint='admin_dt_pdf_single'
)
@login_required(
    ('teacher', 'Resource_Manager')
)
def dt_pdf_single(student_id, dt_number):
    student = db.session.get(
        User,
        student_id
    )
    if not student:
        flash(
            'Student not found.',
            'error'
        )
        return redirect(
            url_for(
                f'{_dt_role_prefix()}_dt'
            )
        )
    rows = []
    for subject in DT_SUBJECTS:
        dt = DiagnosticTest.query.filter_by(
            dt_number=dt_number,
            subject=subject,
            grade=student.grade,
            academic_year=ACADEMIC_YEAR
        ).filter(
            db.or_(
                DiagnosticTest.section == None,
                DiagnosticTest.section == student.section
            )
        ).first()
        if not dt:
            continue
        mark = DTMark.query.filter_by(
            dt_id=dt.id,
            student_id=student_id
        ).first()
        if mark:
            percentage = (
                round(
                    mark.marks_obtained
                    / dt.max_marks
                    * 100,
                    1
                )
                if dt.max_marks
                else 0
            )
            rows.append((
                subject,
                mark.marks_obtained,
                dt.max_marks,
                percentage,
                mark.remarks or ''
            ))
    buffer = io.BytesIO()
    pdf = pdfcanvas.Canvas(
        buffer,
        pagesize=A4
    )
    width, height = A4
    y = _pdf_header(
        pdf,
        f'Diagnostic Test {dt_number} — Result',
        f'{student.grade}'
        f'{(" " + student.section) if student.section else ""}'
        f' · {ACADEMIC_YEAR}'
    )
    pdf.setFont(
        'Helvetica-Bold',
        12
    )
    pdf.drawString(
        40,
        y - 115,
        f'Student: {student.name}'
    )
    pdf.setFont(
        'Helvetica',
        10
    )
    pdf.drawString(
        40,
        y - 132,
        f'Username: {student.username}'
    )
    table_y = y - 165
    pdf.setFillColorRGB(
        0.95,
        0.96,
        0.98
    )
    pdf.rect(
        40,
        table_y - 4,
        width - 80,
        22,
        fill=1,
        stroke=0
    )
    pdf.setFillColorRGB(
        0.2,
        0.25,
        0.35
    )
    pdf.setFont(
        'Helvetica-Bold',
        9
    )
    headers = [
        'Subject',
        'Marks Obtained',
        'Max Marks',
        'Percentage',
        'Remarks'
    ]
    positions = [
        50,
        210,
        310,
        400,
        480
    ]
    for h, x in zip(headers, xpos):
        c.drawString(x, ty + 2, h)
    ty -= 26

    c.setFont('Helvetica', 9)
    total_obt = total_max = 0
    if not rows:
        c.setFillColorRGB(0.6, 0.6, 0.6)
        c.drawString(50, ty + 2, 'No marks recorded for this DT yet.')
        ty -= 20
    for sub, obt, mx, pct, remark in rows:
        c.setFillColorRGB(0, 0, 0)
        c.drawString(50, ty + 2, sub)
        c.drawString(210, ty + 2, str(obt))
        c.drawString(310, ty + 2, str(mx))
        color = (0.06, 0.4, 0.2) if pct >= 80 else (0.57, 0.25, 0.05) if pct >= 60 else (0.6, 0.12, 0.12)
        c.setFillColorRGB(*color)
        c.drawString(400, ty + 2, f'{pct}%')
        c.setFillColorRGB(0, 0, 0)
        c.drawString(480, ty + 2, remark[:22])
        total_obt += obt
        total_max += mx
        ty -= 20

    if rows:
        ty -= 6
        c.setFont('Helvetica-Bold', 10)
        overall_pct = round(total_obt / total_max * 100, 1) if total_max else 0
        c.drawString(50, ty, f'Overall: {total_obt}/{total_max}  ({overall_pct}%)')

    c.setFont('Helvetica-Oblique', 8)
    c.setFillColorRGB(0.5, 0.5, 0.5)
    c.drawString(40, 40, f'Generated on {datetime.utcnow().strftime("%d %b %Y")} · '
                         f'Eastern Public School IBT Portal — for parent sharing')
    c.showPage()
    c.save()
    buf.seek(0)
    fname = f"DT{dt_number}_{student.name.replace(' ', '_')}.pdf"
    return Response(buf.read(), mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={fname}'})


@app.route('/teacher/dt/report/<int:student_id>', endpoint='teacher_dt_report')
@app.route('/admin/dt/report/<int:student_id>', endpoint='admin_dt_report')
@login_required(('teacher', 'Resource_Manager'))
def dt_report(student_id):
    student = db.session.get(User, student_id)
    if not student:
        flash('Student not found.', 'error')
        return redirect(url_for(f'{_dt_role_prefix()}_dt'))
    series = dt_student_series(student_id)

    buf = io.BytesIO()
    c = pdfcanvas.Canvas(buf, pagesize=A4)
    width, height = A4
    y = _pdf_header(c, 'Diagnostic Test — Progress Report',
        f"{student.grade}{(' ' + student.section) if student.section else ''} · {ACADEMIC_YEAR}")
    c.setFont('Helvetica-Bold', 12)
    c.drawString(40, y - 115, f'Student: {student.name}  ({student.username})')

    chart_hex = ['#3b82f6', '#10b981', '#f59e0b', '#8b5cf6']
    chart_colors = [rl_colors.HexColor(h) for h in chart_hex]

    d = Drawing(width - 80, 190)
    lc = HorizontalLineChart()
    lc.x = 30
    lc.y = 20
    lc.width = width - 160
    lc.height = 150
    lc.data = []
    lc.categoryAxis.categoryNames = [f'DT{n}' for n in DT_NUMBERS]
    lc.valueAxis.valueMin = 0
    lc.valueAxis.valueMax = 100
    lc.valueAxis.valueStep = 20
    for i, sub in enumerate(SUBJECTS):
        pts = [(p['pct'] if p['pct'] is not None else 0) for p in series[sub]]
        while len(pts) < 6:
            pts.append(0)
        lc.data.append(pts[:6])
        lc.lines[i].strokeColor = chart_colors[i % len(chart_colors)]
        lc.lines[i].strokeWidth = 2
    d.add(lc)
    renderPDF.draw(d, c, 40, y - 320)

    lx = 40
    c.setFont('Helvetica', 8)
    for i, sub in enumerate(SUBJECTS):
        c.setFillColor(chart_colors[i % len(chart_colors)])
        c.rect(lx, y - 335, 8, 8, fill=1, stroke=0)
        c.setFillColorRGB(0, 0, 0)
        c.drawString(lx + 12, y - 335, sub)
        lx += 115

    ty = y - 370
    c.setFont('Helvetica-Bold', 10)
    c.drawString(40, ty, 'DT-wise Marks (% of max)')
    ty -= 18
    c.setFont('Helvetica-Bold', 8)
    headers = ['Subject'] + [f'DT{n}' for n in DT_NUMBERS] + ['Avg']
    xpos = [45, 135, 180, 225, 270, 315, 360, 415]
    for h, x in zip(headers, xpos):
        c.drawString(x, ty, h)
    ty -= 14

    c.setFont('Helvetica', 8)
    for sub in SUBJECTS:
        c.setFillColorRGB(0, 0, 0)
        c.drawString(45, ty, sub)
        vals = []
        for i in range(6):
            p = series[sub][i]['pct'] if i < len(series[sub]) else None
            c.drawString(xpos[i + 1], ty, f'{p}%' if p is not None else '—')
            if p is not None:
                vals.append(p)
        avg = round(sum(vals) / len(vals), 1) if vals else 0
        c.setFont('Helvetica-Bold', 8)
        c.drawString(xpos[-1], ty, f'{avg}%')
        c.setFont('Helvetica', 8)
        ty -= 14

    c.setFont('Helvetica-Oblique', 8)
    c.setFillColorRGB(0.5, 0.5, 0.5)
    c.drawString(40, 40, f'Generated on {datetime.utcnow().strftime("%d %b %Y")} · '
                         f'Eastern Public School IBT Portal — for parent sharing')
    c.showPage()
    c.save()
    buf.seek(0)
    fname = f"DT_Progress_Report_{student.name.replace(' ', '_')}.pdf"
    return Response(buf.read(), mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={fname}'})


# ── STUDENT DIAGNOSTIC PROGRESS ────────────────────────────────────────────────

@app.route('/student/diagnostics')
@login_required('student')
def student_diagnostics():
    student = db.session.get(User, session['user_id'])
    series = dt_student_series(student.id)
    insights = dt_student_insights(series)
    weak_subjects = [item['subject'] for item in insights
                     if item['average'] is not None and item['average'] < 80]
    practice_tests = MockTest.query.filter(
        MockTest.status == 'active',
        db.or_(MockTest.grade == student.grade, MockTest.grade == 'All Grades')
    ).order_by(MockTest.subject, MockTest.name).all()
    return render_template('student/diagnostics.html',
        student=student, series=series, insights=insights,
        subjects=SUBJECTS, dt_numbers=DT_NUMBERS,
        weak_subjects=weak_subjects, practice_tests=practice_tests,
        academic_year=ACADEMIC_YEAR)


# Initialise after every model—including diagnostic models—has been declared.
with app.app_context():
    db.create_all()
    seed_db()


if __name__ == '__main__':
    print("\n🎓 Eastern Public School — IBT Portal")
    print("   Admin: Organizer / bk*123\n")
    app.run(debug=True, host='0.0.0.0', port=5000)
